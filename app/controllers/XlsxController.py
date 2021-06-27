from openpyxl import load_workbook

class XlsxController():
    file = ''
    letter_snils = ''
    letter_status = ''
    row_start = 0
    mass = []

    def __init__(self, file, row_start, letter_snils, letter_status):
        self.snilses = []
        self.file = file
        self.row_start = row_start
        self.letter_snils = letter_snils
        self.letter_status = letter_status

    def snilsToNumber(self, number):
        number = str(number)
        number = number.replace(' ', '').replace('-', '')
        if len(number) == 11:
            return number
        else:
            return False

    def numberToSnils(self, number):
        if len(number) == 11:
            n1 = number[0:3]
            n2 = number[3:6]
            n3 = number[6:9]
            n4 = number[9:11]
            snils = n1 + '-' + n2 + '-' + n3 + ' ' + n4
            return snils
        else:
            return False

    def getData(self):
        file = self.file
        wb = load_workbook(filename=file)
        ws = wb.active
        max_row = ws.max_row
        for i in range(self.row_start, max_row + 1):

            cell_coord_snils = self.letter_snils + str(i)
            snils = ws[cell_coord_snils].value

            cell_coord_status = self.letter_status + str(i)
            status = ws[cell_coord_status].value

            num_snils = self.snilsToNumber(snils)
            snils = self.numberToSnils(num_snils)

            if snils:
                row = {"snils": snils, "status": status}
                self.mass.append(row)
        return self.mass
